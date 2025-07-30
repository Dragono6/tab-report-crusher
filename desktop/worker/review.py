"""
Main Python worker for TAB Report Crusher.

This script will be bundled into a PyInstaller executable and called by the Tauri shell.
It handles file parsing, AI review, and PDF annotation.
"""

import sys
import json
import openpyxl
import fitz  # PyMuPDF
import pdfplumber
import re
from pathlib import Path

from database import create_schema
from ai_gateway import chunk_data, run_ai_review
from annotator import add_annotations_to_pdf

# TODO: Import AI libraries and other dependencies

def extract_data_from_pdf(file_path: str) -> list:
    """Extracts tables and text from a PDF file."""
    data = []
    with pdfplumber.open(file_path) as pdf:
        for i, page in enumerate(pdf.pages):
            # Extract tables
            tables = page.extract_tables()
            for table in tables:
                data.append({"type": "table", "page": i + 1, "content": table})
            
            # Extract text, preserving some structure
            text = page.extract_text()
            data.append({"type": "text", "page": i + 1, "content": text})
            
    print(f"Extracted {len(data)} data chunks from PDF.", file=sys.stderr)
    return data

def extract_data_from_excel(file_path: str) -> list:
    """Extracts data from all sheets in an Excel file."""
    data = []
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []
        for row in sheet.iter_rows():
            sheet_data.append([cell.value for cell in row])
        data.append({"type": "sheet", "name": sheet_name, "content": sheet_data})
    
    print(f"Extracted {len(data)} sheets from Excel file.", file=sys.stderr)
    return data

def main():
    """
    Main function to process the report.
    Takes a file path as a command-line argument.
    """
    # Initialize the local database schema
    create_schema()
    
    # Get arguments from command line
    if len(sys.argv) > 3:
        file_path = sys.argv[1]
        api_key = sys.argv[2]
        selected_model = sys.argv[3]
    else:
        print("Error: Missing arguments (file_path, api_key, model_name).", file=sys.stderr)
        sys.exit(1)
    
    print(f"Processing file: {file_path}", file=sys.stderr)
    
    file_extension = Path(file_path).suffix.lower()
    
    extracted_data = []
    if file_extension == '.pdf':
        extracted_data = extract_data_from_pdf(file_path)
    elif file_extension in ['.xlsx', '.xls']:
        extracted_data = extract_data_from_excel(file_path)
    else:
        print(f"Unsupported file type: {file_extension}", file=sys.stderr)
        return

    # Chunk the data and run the AI review
    data_chunks = chunk_data(extracted_data, selected_model)
    # TODO: Load profile from DB instead of using placeholder
    active_profile = {"name": "Manager Default"}
    findings = run_ai_review(api_key, selected_model, data_chunks, active_profile)

    # TODO: Implement rule engine logic
    
    # Annotate the PDF with the findings
    if file_extension == '.pdf':
        output_path = Path(file_path).with_name(f"{Path(file_path).stem}_review.pdf")
        add_annotations_to_pdf(file_path, findings, str(output_path))
    
    # TODO: Return results as JSON to stdout
    result = {"status": "success", "findings": findings}
    print(json.dumps(result))

if __name__ == "__main__":
    main() 